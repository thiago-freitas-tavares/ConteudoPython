# Como Processar Planilhas

# Uma aplicação python consegue processar milhares de planilhas em menos de 1s, o que levaria horas ou dias manualmente.

# Faremos um programa python para acessar a planilha transactions.
# Reduzir em 10% os valores da coluna C e incluir um gráfico.

import openpyxl as xl       # demos o pseudônimo xl para o pacote openpyxl para o código ficar mais clean.
from openpyxl.chart import BarChart, Reference


def process_workbook(file_name):             # -> Colocamos tudo dentro de uma função.
#   wb = xl.load_workbook('transactions.xlsx') -> Dessa forma o programa funciona apenas para este arquivo.
    wb = xl.load_workbook(file_name)         # -> Dessa forma funciona para vários arquivos (com a função).
    sheet = wb['Sheet1']
#   cell = sheet['a1']          # Duas formas de buscar uma célula específica.
    cell = sheet.cell(1, 1)     # Duas formas de buscar uma célula específica.
    print(cell.value)

# range() começa em 2, pois a primeira linha é cabeçalho e a contagem de linhas de sheet.cell começa em 1.
# range() não inclui o segundo argumento, por isso somamos + 1 ao máximo de linhas da planilha.
    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price

    # Agora faremos um gráfico à partir dos novos dados (células D2, D3 e D4)

    values = Reference(sheet,           # A classe Reference seleciona um range de valores.
              min_row=2,
              max_row=sheet.max_row,
              min_col=4,
              max_col=4)

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')        # O segundo argumento é a célula da parte superior esquerda do gráfico

#   wb.save('transactions2.xlsx')       # Dessa forma cria um novo arquivo com o nome especificado.
    wb.save(file_name)                  # Dessa forma substitui o mesmo arquivo.

# Daria para fazer um programa para encontrar todos os arquivos dentro de um repositório
# Passar o nome de todos os arquivos para a função process_workbook
# Teríamos todos arquivos do repositório atualizados de acordo com o código da função em fração de segundo.
